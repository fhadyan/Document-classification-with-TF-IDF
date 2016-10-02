import os
import xml.etree.ElementTree as ET
import nltk
from nltk import FreqDist
import re
from nltk.corpus import stopwords
from nltk.stem.lancaster import LancasterStemmer
from nltk.stem.porter import PorterStemmer
import Preprocessing
import tfidf
import numpy as np
from collections import defaultdict
from math import log10
from openpyxl import Workbook
from openpyxl.utils import (_get_column_letter)


class Weighting:
    def __init__(self):
        self.sourcePath = "output/training/all/"
        self.outPath = "output/training/tfidf/"
        self.TestSourcePath = "output/testing/all/"
        self.TestOutPath = "output/testing/tfidf/"
        if os.listdir(self.sourcePath)=="":
            d = Preprocessing()
            d.output()

    def tfidf(self):
        #table = tfidf.tfidf()
        allString = []
        idf = []
        docCount=0
        tf=[]
        for filename in os.listdir(self.sourcePath):
            docCount+=1
            if not filename.endswith('.txt'): continue
            fullname = os.path.join(self.sourcePath, filename)

            file = open(fullname, 'r+')
            out = open(self.outPath + filename, 'wb+')
            strings = file.readlines()

            # temptf = defaultdict(int)
            # #table.addDocument(filename, strings)
            idf.extend(set(strings))
            for string in strings:
                allString.append(string[:-1])
            #     temptf[string[:-1]]+=1
            # for item in strings:
            #     temptf[item[:-1]]=1+log10(temptf[item[:-1]])
            # tf.append(temptf)
        idfres = defaultdict(int)
        for item in idf:
            idfres[item[:-1]]+=1
        for item in set(idf):
            idfres[item[:-1]]= log10(docCount/idfres[item[:-1]])
        # print(idfres)

        #tf test set
        tf=[]
        trainingNames=[]
        for filename in os.listdir(self.sourcePath):
            if not filename.endswith('.txt'): continue
            fullname = os.path.join(self.sourcePath, filename)

            trainingNames.append(filename)
            file = open(fullname, 'r+')
            strings = file.readlines()

            temptf = defaultdict(int)
            for string in strings:
                allString.append(string[:-1])
                temptf[string[:-1]]+=1
            for item in set(strings):
                temptf[item[:-1]]=1+log10(temptf[item[:-1]])
            tf.append(temptf)
            #print(temptf)

        # for item in tf:
        #     print(item)

        tfidfTraining = []
        for item1 in tf:
            temp = defaultdict(int)
            for item2 in idfres:
                temp[item2] = idfres[item2]*item1[item2]
            tfidfTraining.append(temp)
        # for item in tfidfTraining:
        #     print(item)


        tf=[]
        testingNames=[]
        for filename in os.listdir(self.TestSourcePath):
            if not filename.endswith('.txt'): continue
            fullname = os.path.join(self.TestSourcePath, filename)

            testingNames.append(filename)
            file = open(fullname, 'r+')
            out = open(self.outPath + filename, 'wb+')
            strings = file.readlines()

            temptf = defaultdict(int)
            for string in strings:
                allString.append(string[:-1])
                temptf[string[:-1]]+=1
            for item in strings:
                temptf[item[:-1]]=1+log10(temptf[item[:-1]])
            tf.append(temptf)

        tfidfTesting = []
        for item1 in tf:
            temp = defaultdict(int)
            for item2 in idfres:
                temp[item2] = idfres[item2]*item1[item2]
            tfidfTesting.append(temp)


        names=[]
        names.append('filename')
        for item in idfres:
            names.append(item)

        wb = Workbook()
        ws = wb.active
        ws.append(names)
        i=2
        k=0
        for item in tfidfTraining:
            j=2
            ws[_get_column_letter(j-1)+ str(i)] = trainingNames[k]
            for item2 in idfres:
                ws[_get_column_letter(j)+ str(i)] = item[item2]
                j+=1
            i+=1
            k+=1
        wb.save(self.TestOutPath+"training.xlsx")
        # print(len(tfidfTraining))
        # print(len(trainingNames))

        wb = Workbook()
        ws = wb.active
        ws.append(names)
        i=2
        k=0
        for item in tfidfTesting:
            j=2
            ws[_get_column_letter(j-1)+ str(i)] = testingNames[k]
            for item2 in idfres:
                ws[_get_column_letter(j)+ str(i)] = item[item2]
                j+=1
            i+=1
            k+=1
        wb.save(self.TestOutPath+"testing.xlsx")

        # print(len(tfidfTraining))
        # print(len(tfidfTesting))

w = Weighting()
w.tfidf();
